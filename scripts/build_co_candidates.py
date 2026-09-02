#!/usr/bin/env python3
"""Parse official Colorado SOS 2026 primary + general candidate Excel lists."""

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

UA = "WeThePeople-CivicBot/1.0"
RETRIEVED = "2026-09-02T11:12:00Z"
PRIMARY_URL = "https://www.sos.state.co.us/pubs/elections/vote/files/2026/2026PrimaryCandidateListOfficial.xlsx"
GENERAL_URL = "https://www.sos.state.co.us/pubs/elections/vote/files/2026/2026GeneralCandidateListUnofficial.xlsx"
ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "public" / "data" / "co"
STUB = ROOT / "public" / "data" / "co.json"
CACHE = Path("/tmp/co-sos")

LISTS = (
    {
        "url": PRIMARY_URL,
        "filename": "2026PrimaryCandidateListOfficial.xlsx",
        "list_kind": "primary_official",
    },
    {
        "url": GENERAL_URL,
        "filename": "2026GeneralCandidateListUnofficial.xlsx",
        "list_kind": "general_unofficial",
    },
)


def district_value(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    if isinstance(raw, int):
        return str(raw)
    text = str(raw).strip()
    if not text or text.casefold() in {"statewide", "state", "none"}:
        return None
    try:
        num = float(text)
        if num == int(num):
            return str(int(num))
    except ValueError:
        pass
    return text


def contest_key(office: str, district: str | None) -> str:
    return f"CO|{office}|{district or ''}|"


def is_end_marker(name: str) -> bool:
    folded = " ".join(name.casefold().split())
    return folded in {"end of data", "end of data."} or folded.startswith("end of data")


def write_in_value(raw) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    return text


def parse_workbook(path: Path, list_kind: str, source_url: str) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise SystemExit(f"empty workbook {path}")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    expected = ["Candidate Name", "Office", "District", "Party", "Write In?"]
    if header[:5] != expected:
        raise SystemExit(f"unexpected header in {path}: {header[:5]}")
    out: list[dict] = []
    for raw in rows[1:]:
        name = str(raw[0]).strip() if raw[0] is not None else ""
        if not name or is_end_marker(name):
            continue
        office = str(raw[1]).strip() if raw[1] is not None else ""
        if not office:
            raise SystemExit(f"missing office for filed name {name!r} in {path}")
        district = district_value(raw[2])
        party = str(raw[3]).strip() if raw[3] is not None else ""
        rec = {
            "state": "CO",
            "contest_key": contest_key(office, district),
            "office": office,
            "district": district,
            "party": party or None,
            "candidate_name": name,
            "write_in": write_in_value(raw[4]),
            "list_kind": list_kind,
            "source_url": source_url,
            "retrieved_at": RETRIEVED,
        }
        out.append(rec)
    return out


def fetch_if_needed() -> None:
    import urllib.request

    CACHE.mkdir(parents=True, exist_ok=True)
    for item in LISTS:
        dest = CACHE / item["filename"]
        if dest.exists() and dest.stat().st_size > 1000:
            continue
        req = urllib.request.Request(item["url"], headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=90) as resp:
            dest.write_bytes(resp.read())


def summarize(rows: list[dict]) -> dict:
    general = [r for r in rows if r["list_kind"] == "general_unofficial"]
    primary = [r for r in rows if r["list_kind"] == "primary_official"]
    gen_keys = {r["contest_key"] for r in general}

    def office_count(kind: str, *names: str) -> int:
        return sum(1 for r in general if r.get("office") in names)

    house = [r for r in general if r.get("office") == "US House of Representatives"]
    return {
        "row_count": len(rows),
        "contest_key_count": len(gen_keys),
        "primary_official_rows": len(primary),
        "general_unofficial_rows": len(general),
        "us_senate": office_count("US Senate", "US Senate"),
        "us_house": len(house),
        "us_house_districts": sorted(
            {int(r["district"]) for r in house if str(r.get("district") or "").isdigit()}
        ),
        "governor": office_count("Governor", "Governor"),
        "state_senate": office_count("State Senate", "State Senate"),
        "state_house": office_count("State House of Representatives", "State House of Representatives"),
        "by_list_kind": dict(Counter(r["list_kind"] for r in rows)),
        "source_urls": [PRIMARY_URL, GENERAL_URL],
        "retrieved_at": RETRIEVED,
        "note": (
            "Filed names from official Colorado SOS Excel. Primary official and general unofficial "
            "lists are kept separate (list_kind). People are not merged across lists. "
            "contest_key_count is unique general_unofficial contests. End-of-Data rows skipped. "
            "No Ballotpedia. Streets omitted."
        ),
    }


def merge_stub() -> None:
    stub = json.loads(STUB.read_text(encoding="utf-8"))
    donors = ((stub.get("state_filings") or {}).get("donors") or {}).copy()
    election = stub.setdefault("election", {})
    election["note"] = (
        "Official SOS primary/general candidate lists, Clerk/LIS federal votes, "
        "and TRACER Schedule A donors. Donor lists are not sold."
    )
    stub["candidates_path"] = "/data/co/candidates.json"
    stub["candidate_summary_path"] = "/data/co/candidate-summary.json"
    sources = stub.setdefault("sources", [])
    have = {s.get("url") for s in sources if isinstance(s, dict)}
    for url, note in (
        (PRIMARY_URL, "CO SOS 2026 primary official candidate Excel"),
        (GENERAL_URL, "CO SOS 2026 general unofficial candidate Excel"),
    ):
        if url not in have:
            sources.append({"url": url, "retrieved_at": RETRIEVED, "note": note})
    if donors:
        stub["state_filings"]["donors"] = donors
    STUB.write_text(json.dumps(stub, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("merged co.json donors.status", donors.get("status"))


def main() -> int:
    fetch_if_needed()
    rows: list[dict] = []
    for item in LISTS:
        parsed = parse_workbook(CACHE / item["filename"], item["list_kind"], item["url"])
        print(f"{item['list_kind']} rows={len(parsed)}")
        rows.extend(parsed)
    if len(rows) != 661:
        raise SystemExit(f"expected 661 official rows, got {len(rows)}")
    gen_keys = {r["contest_key"] for r in rows if r["list_kind"] == "general_unofficial"}
    if len(gen_keys) != 166:
        raise SystemExit(f"expected 166 general contest_keys, got {len(gen_keys)}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "candidates.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    summary = summarize(rows)
    (OUT_DIR / "candidate-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    merge_stub()
    print(f"wrote {len(rows)} candidates, general contest_keys={len(gen_keys)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
