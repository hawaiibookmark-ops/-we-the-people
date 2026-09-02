#!/usr/bin/env python3
"""Parse official Oregon ORESTAR 2026 primary + general candidate filing exports."""

from __future__ import annotations

import http.cookiejar
import json
import re
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

UA = "WeThePeople-CivicBot/1.0"
RETRIEVED = "2026-09-02T11:20:30Z"
SEARCH_URL = "https://secure.sos.state.or.us/orestar/CFSearchPage.do"
ELECTIONS_URL = "https://secure.sos.state.or.us/orestar/ajaxdataserver/getCandidateElectionByYear?elecYear=2026"
CSRF_URL = "https://secure.sos.state.or.us/orestar/JavaScriptServlet"
ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "public" / "data" / "or"
STUB = ROOT / "public" / "data" / "or.json"
CACHE = Path("/tmp/or-orestar")

PRIMARY_ELECTION_ID = "1451"
GENERAL_ELECTION_ID = "1453"

LISTS = (
    {
        "election_id": PRIMARY_ELECTION_ID,
        "election_txt": "2026 Primary Election",
        "list_kind": "primary_candidate_filing",
        "filename": "2026-primary.xlsx",
    },
    {
        "election_id": GENERAL_ELECTION_ID,
        "election_txt": "2026 General Election",
        "list_kind": "general_candidate_filing",
        "filename": "2026-general.xlsx",
    },
)


def contest_key(office: str, district: str | None) -> str:
    return f"OR|{office}|{district or ''}|"


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_workbook(path: Path, list_kind: str, election_id: str, election_txt: str) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise SystemExit(f"empty workbook {path}")
    header = [cell_text(c) for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    required = (
        "Election Txt",
        "Election Year",
        "Office Group",
        "Office",
        "Candidate Office",
        "Candidate File RSN",
        "Filetype Descr",
        "Party Descr",
        "Cand Ballot Name Txt",
    )
    missing = [name for name in required if name not in idx]
    if missing:
        raise SystemExit(f"unexpected ORESTAR header in {path}: missing {missing}")
    out: list[dict] = []
    for raw in rows[1:]:
        name = cell_text(raw[idx["Cand Ballot Name Txt"]])
        office = cell_text(raw[idx["Office Group"]])
        district = cell_text(raw[idx["Office"]]) or None
        if not name or not office:
            raise SystemExit(f"missing filed name/office in {path}: {name!r} {office!r}")
        election = cell_text(raw[idx["Election Txt"]]) or election_txt
        year = cell_text(raw[idx["Election Year"]]) or "2026"
        rec = {
            "state": "OR",
            "contest_key": contest_key(office, district),
            "office": office,
            "district": district,
            "candidate_office": cell_text(raw[idx["Candidate Office"]]) or None,
            "party": cell_text(raw[idx["Party Descr"]]) or None,
            "candidate_name": name,
            "list_kind": list_kind,
            "election": election,
            "election_year": year,
            "election_id": election_id,
            "filing_type": cell_text(raw[idx["Filetype Descr"]]) or None,
            "candidate_file_rsn": cell_text(raw[idx["Candidate File RSN"]]) or None,
            "source_url": SEARCH_URL,
            "retrieved_at": RETRIEVED,
        }
        out.append(rec)
    return out


def _request(opener: urllib.request.OpenerDirector, url: str, data=None, headers=None, method=None):
    h = {
        "User-Agent": UA,
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.9",
    }
    if headers:
        h.update(headers)
    req = urllib.request.Request(url, data=data, headers=h, method=method)
    with opener.open(req, timeout=90) as resp:
        return resp.geturl(), resp.headers, resp.read()


def fetch_one(election_id: str, dest: Path) -> None:
    if dest.exists() and dest.stat().st_size > 1000:
        return
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    _request(opener, SEARCH_URL)
    jsid = next((c.value for c in cj if c.name == "JSESSIONID_ORESTAR"), None)
    if not jsid:
        raise SystemExit("ORESTAR session cookie missing")
    _, _, tok = _request(
        opener,
        CSRF_URL,
        data=b"",
        headers={
            "FETCH-CSRF-TOKEN": "1",
            "Referer": SEARCH_URL,
            "Origin": "https://secure.sos.state.or.us",
        },
        method="POST",
    )
    token_name, token_value = tok.decode("utf-8").split(":", 1)
    fields = {
        "cfSearchButtonName": "",
        "cfName": "",
        "cfyearActive": "2026",
        "cfElection": election_id,
        "cfOffice": "",
        "cfOfficeGrp": "",
        "cfPartyAffiliation": "",
        "cfFilingType": "",
        "cfFilingFromDate": "",
        "cfFilingToDate": "",
        "cfWithDrawFromDate": "",
        "cfWithDrawToDate": "",
        token_name: token_value,
    }
    action = f"https://secure.sos.state.or.us/orestar/cfFilings.do;JSESSIONID_ORESTAR={jsid}"
    _, _, results = _request(
        opener,
        action,
        data=urllib.parse.urlencode(fields).encode(),
        headers={
            "Referer": SEARCH_URL,
            "Origin": "https://secure.sos.state.or.us",
            "Content-Type": "application/x-www-form-urlencoded",
            token_name: token_value,
        },
    )
    text = results.decode("utf-8", "replace")
    m = re.search(r"XcelCFSearch[^\"']+", text)
    if not m:
        raise SystemExit(f"ORESTAR Excel export link missing for election {election_id}")
    export_url = "https://secure.sos.state.or.us/orestar/" + m.group(0)
    _, _, xls = _request(
        opener,
        export_url,
        headers={"Referer": action, token_name: token_value},
    )
    if xls[:2] != b"PK":
        raise SystemExit(f"ORESTAR export was not Excel for election {election_id}")
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(xls)
    print(f"fetched {dest} bytes={len(xls)} election={election_id}", flush=True)


def fetch_if_needed() -> None:
    CACHE.mkdir(parents=True, exist_ok=True)
    for item in LISTS:
        fetch_one(item["election_id"], CACHE / item["filename"])


def district_number(district: str | None) -> int | None:
    if not district:
        return None
    m = re.match(r"(\d+)", district)
    return int(m.group(1)) if m else None


def summarize(rows: list[dict]) -> dict:
    primary = [r for r in rows if r["list_kind"] == "primary_candidate_filing"]
    general = [r for r in rows if r["list_kind"] == "general_candidate_filing"]
    keys = {r["contest_key"] for r in rows}

    def office_count(kind_rows: list[dict], *names: str) -> int:
        return sum(1 for r in kind_rows if r.get("office") in names)

    house = [r for r in general if r.get("office") == "US Representative"]
    return {
        "row_count": len(rows),
        "contest_key_count": len(keys),
        "primary_candidate_filing_rows": len(primary),
        "general_candidate_filing_rows": len(general),
        "us_senate": office_count(general, "US Senator"),
        "us_house": len(house),
        "us_house_districts": sorted(
            {district_number(r.get("district")) for r in house if district_number(r.get("district"))}
        ),
        "governor": office_count(general, "Governor"),
        "state_senate": office_count(general, "State Senator"),
        "state_house": office_count(general, "State Representative"),
        "by_list_kind": dict(Counter(r["list_kind"] for r in rows)),
        "by_office": dict(Counter(r["office"] for r in rows)),
        "source_urls": [SEARCH_URL, ELECTIONS_URL],
        "election_ids": {
            "primary": PRIMARY_ELECTION_ID,
            "general": GENERAL_ELECTION_ID,
        },
        "retrieved_at": RETRIEVED,
        "note": (
            "Filed ballot names from official ORESTAR Candidate Filing Search Excel "
            "(XcelCFSearch, year=2026, elections 1451 primary and 1453 general). "
            "list_kind and election fields are kept separate. People are not merged "
            "across lists or fusion-party rows. Streets, phones, and emails are omitted. "
            "No Ballotpedia. Names are not invented."
        ),
    }


def write_stub() -> None:
    if STUB.exists():
        stub = json.loads(STUB.read_text(encoding="utf-8"))
    else:
        stub = {}
    donors = ((stub.get("state_filings") or {}).get("donors") or {}).copy()
    if not donors:
        donors = {
            "status": "pending",
            "reason": (
                "Oregon has no free statewide ORESTAR/Schedule A bulk dump for this populate. "
                "Donor names are not invented."
            ),
            "do_not_sell_donor_lists": True,
        }
    stub["election"] = {
        "jurisdiction": "Oregon",
        "state_code": "OR",
        "general_date": None,
        "note": (
            "Official ORESTAR 2026 primary/general candidate filings and Clerk/LIS federal votes. "
            "Oregon has no free statewide donor bulk dump; donor names are not invented. "
            "Donor lists are not sold."
        ),
    }
    stub["state_filings"] = {
        "wired": True,
        "orestar_public": SEARCH_URL,
        "donors": donors,
    }
    stub.setdefault("nominees", {})
    stub.setdefault("geo_by_zip", {})
    stub["candidates_path"] = "/data/or/candidates.json"
    stub["candidate_summary_path"] = "/data/or/candidate-summary.json"
    sources = stub.setdefault("sources", [])
    have = {s.get("url") for s in sources if isinstance(s, dict)}
    extra = [
        {
            "url": SEARCH_URL,
            "retrieved_at": RETRIEVED,
            "note": "ORESTAR Candidate Filing Search year=2026 (primary 1451, general 1453)",
        },
        {
            "url": ELECTIONS_URL,
            "retrieved_at": RETRIEVED,
            "note": "ORESTAR official 2026 election ids",
        },
    ]
    for src in extra:
        if src["url"] not in have:
            sources.append(src)
    STUB.parent.mkdir(parents=True, exist_ok=True)
    STUB.write_text(json.dumps(stub, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("wrote stub", STUB, "donors.status", donors.get("status"), flush=True)


def main() -> int:
    fetch_if_needed()
    rows: list[dict] = []
    for item in LISTS:
        parsed = parse_workbook(
            CACHE / item["filename"],
            item["list_kind"],
            item["election_id"],
            item["election_txt"],
        )
        print(f"{item['list_kind']} rows={len(parsed)}", flush=True)
        rows.extend(parsed)
    if len(rows) != 604:
        raise SystemExit(f"expected 604 official filing rows, got {len(rows)}")
    primary_n = sum(1 for r in rows if r["list_kind"] == "primary_candidate_filing")
    general_n = sum(1 for r in rows if r["list_kind"] == "general_candidate_filing")
    if primary_n != 343 or general_n != 261:
        raise SystemExit(f"expected 343/261 primary/general, got {primary_n}/{general_n}")
    keys = {r["contest_key"] for r in rows}
    if len(keys) != 166:
        raise SystemExit(f"expected 166 contest_keys, got {len(keys)}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "candidates.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    summary = summarize(rows)
    (OUT_DIR / "candidate-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    write_stub()
    print(f"wrote {len(rows)} candidates, contest_keys={len(keys)}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
